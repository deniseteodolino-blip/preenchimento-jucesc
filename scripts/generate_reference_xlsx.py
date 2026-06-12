#!/usr/bin/env python3
"""
Gera planilha Excel de referência a partir do parse de uma minuta.

Útil para:
- Backup manual: se a automação falhar em algum sócio, o usuário tem todos os
  códigos JUCESC/JUCEB prontos para preenchimento manual
- Tracking: marcar em verde sócios já preenchidos
- Sessões longas: continuar em outra conversa sem perder contexto

Uso:
    python generate_reference_xlsx.py <socios.json> <output.xlsx> [<capital.json>]
"""
import json
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

EC_LABELS = {
    '492': 'SOLTEIRO',
    '493': 'CASADO',
    '494': 'DESQUITADO',
    '495': 'SEPARADO',
    '496': 'VIÚVO',
    '6574': 'DIVORCIADO',
    '6863': 'SOLTEIRO EM UNIÃO ESTÁVEL',
    '6864': 'DIVORCIADO EM UNIÃO ESTÁVEL',
    '6865': 'VIÚVO EM UNIÃO ESTÁVEL',
    '7111': 'SOLTEIRO EM UNIÃO ESTÁVEL (JUCEB)',
    '7129': 'DIVORCIADO EM UNIÃO ESTÁVEL (JUCEB)',
    '7130': 'VIÚVO EM UNIÃO ESTÁVEL (JUCEB)',
}

REG_LABELS = {
    '991': 'COMUNHÃO PARCIAL',
    '992': 'COMUNHÃO UNIVERSAL',
    '993': 'PARTICIPAÇÃO FINAL NOS AQUESTOS',
    '994': 'SEPARAÇÃO DE BENS',
}


def main():
    if len(sys.argv) < 3:
        print('Uso: python generate_reference_xlsx.py <socios.json> <output.xlsx> [<capital.json>]', file=sys.stderr)
        sys.exit(1)
    socios = json.load(open(sys.argv[1], encoding='utf-8'))
    out = sys.argv[2]
    capital = None
    if len(sys.argv) > 3:
        capital_data = json.load(open(sys.argv[3], encoding='utf-8'))
        capital = {c['cpf']: c for c in capital_data}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'QSA'

    headers = [
        '#', 'Tipo', 'CPF/CNPJ', 'Nome', 'Sexo (heur.)',
        'EC JUCESC', 'EC JUCEB', 'EC Texto',
        'Regime', 'Regime Texto',
        'Tipo Doc (default 2076)',
        'RG', 'Órgão', 'UF Órgão',
        'Profissão', 'DDD', 'Telefone', 'Email',
        'CEP', 'Data Nasc', 'NIRE (PJ)', 'Representante (PJ)',
        'Status'
    ]
    if capital:
        headers.extend(['Quotas', 'Capital Integralizado', 'Capital A Integralizar'])

    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for s in socios:
        ec = s.get('ec_jucesc') or ''
        eb = s.get('ec_juceb') or ''
        reg = s.get('regime_jucesc') or ''
        row = [
            s.get('num'),
            s.get('tipo'),
            s.get('cpf') or s.get('cnpj') or '',
            s.get('nome', ''),
            s.get('sexo_heuristico', '') if s.get('tipo') == 'PF' else '',
            ec, eb, s.get('estado_civil_raw', ''),
            reg, REG_LABELS.get(reg, ''),
            '2076',
            s.get('rg', '') or '',
            s.get('orgao_expedidor', '') or '',
            s.get('uf_orgao', '') or '',
            s.get('profissao', '') or '',
            s.get('ddd', ''),
            s.get('telefone', ''),
            s.get('email', ''),
            s.get('cep', ''),
            s.get('data_nascimento', '09/08/1980'),
            s.get('nire', '') if s.get('tipo') == 'PJ' else '',
            s.get('representante_nome', '') if s.get('tipo') == 'PJ' else '',
            'Pendente',
        ]
        if capital:
            cap = capital.get(s.get('cpf') or s.get('cnpj') or '')
            if cap:
                row.extend([cap['quotas'], cap['integralizado'], cap['a_integralizar']])
            else:
                row.extend(['', '', ''])
        ws.append(row)

    # Auto-tamanho de colunas
    for col in ws.columns:
        maxlen = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(35, maxlen + 2)
    ws.freeze_panes = 'A2'

    # Aba de Códigos
    ws2 = wb.create_sheet('Códigos')
    ws2.append(['Sistema', 'Campo', 'Código', 'Valor'])
    for c in ws2[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    for code, label in EC_LABELS.items():
        if code in ('7111', '7129', '7130'):
            ws2.append(['JUCEB', 'Estado Civil', int(code), label])
        else:
            ws2.append(['JUCESC', 'Estado Civil', int(code), label])
    for code, label in REG_LABELS.items():
        ws2.append(['Ambos', 'Regime', int(code), label])
    ws2.append(['Ambos', 'Tipo Doc', 2076, 'CARTEIRA DE IDENTIDADE'])
    ws2.append(['Ambos', 'Tipo Doc', 516, 'CARTEIRA NACIONAL DE HABILITAÇÃO'])
    ws2.append(['Ambos', 'Tipo Doc', 6562, 'CARTEIRA DE IDENTIDADE PROFISSIONAL'])
    ws2.append(['Ambos', 'Tipo Doc', 6573, 'REGISTRO NACIONAL DE ESTRANGEIRO'])
    ws2.append(['Ambos', 'Tipo Doc', 7038, 'REGISTRO NACIONAL MIGRATÓRIO'])
    ws2.append(['Ambos', 'Tipo Assistido (Rep)', 304, 'REPRESENTANTE (default p/ PJ)'])

    for col in ws2.columns:
        maxlen = max(len(str(c.value or '')) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = min(45, maxlen + 2)

    wb.save(out)
    print(f'→ Planilha gerada: {out}')
    print(f'   {len(socios)} sócios | {ws2.max_row - 1} códigos de referência')


if __name__ == '__main__':
    main()
